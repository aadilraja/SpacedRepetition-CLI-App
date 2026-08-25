import importlib.util
import json
import tempfile
import unittest
from datetime import date, timedelta
from pathlib import Path

from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from typer.testing import CliRunner


MODULE_PATH = Path(__file__).with_name("main.py")

spec = importlib.util.spec_from_file_location("main", MODULE_PATH)
revise = importlib.util.module_from_spec(spec)
spec.loader.exec_module(revise)


class RevisionTrackerTests(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.engine = create_engine("sqlite:///:memory:", echo=False)
        revise.engine = cls.engine
        revise.Base.metadata.create_all(cls.engine)

    def setUp(self):
        with Session(self.engine) as session:
            for table in reversed(revise.Base.metadata.sorted_tables):
                session.execute(table.delete())
            session.commit()

    def make_tracker(self, content):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "tracker.md"
        path.write_text(content, encoding="utf-8")
        return path

    def make_config(self, path):
        return {"name": "Test User", "tracker_path": str(path), "last_scan": None}

    def get_titles(self):
        with Session(self.engine) as session:
            topics = session.query(revise.Coverage).order_by(revise.Coverage.id).all()
            return [topic.title for topic in topics]

    def add_coverage(
        self,
        title,
        covered_date=None,
        next_review=None,
        interval_days=2,
        stage=0,
        status="active",
        keywords=None,
        source_key=None
    ):
        """
        Insert a Coverage row directly, bypassing the Markdown parser.

        Used by tests that only care about database-level behavior
        (review scheduling, filtering, removal) rather than parsing.
        """

        import json

        covered_date = covered_date or date.today()
        next_review = next_review or covered_date

        with Session(self.engine) as session:
            coverage = revise.Coverage(
                title=title,
                covered_date=covered_date,
                next_review=next_review,
                interval_days=interval_days,
                stage=stage,
                status=status,
                keywords=json.dumps(keywords or []),
                source_file="manual-test-entry",
                source_key=source_key or f"manual::{title}::{covered_date.isoformat()}"
            )
            session.add(coverage)
            session.commit()
            session.refresh(coverage)
            return coverage.id

    def test_scan_only_reads_top_date_and_only_x(self):
        tracker = self.make_tracker(
            "# 12/06/2026\n\n- [ ] topic 1\n- [x] topic 2\n\n# 13/06/2026\n\n- [ ] topic 3\n- [x] topic 4\n"
        )
        revise.scan_tracker(self.make_config(tracker))
        self.assertEqual(self.get_titles(), ["topic 2"])

    def test_unchecked_topic_is_not_added(self):
        tracker = self.make_tracker("# 12/06/2026\n\n- [ ] topic 1\n- [ ] topic 2\n")
        revise.scan_tracker(self.make_config(tracker))
        self.assertEqual(self.get_titles(), [])

    def test_x_and_X_are_checked(self):
        tracker = self.make_tracker("# 12/06/2026\n\n- [x] lowercase x\n- [X] uppercase X\n- [ ] unchecked\n")
        revise.scan_tracker(self.make_config(tracker))
        self.assertEqual(self.get_titles(), ["lowercase x", "uppercase X"])

    def test_scanning_twice_does_not_duplicate(self):
        tracker = self.make_tracker("# 12/06/2026\n\n- [x] topic 1\n")
        config = self.make_config(tracker)
        revise.scan_tracker(config)
        revise.scan_tracker(config)
        self.assertEqual(self.get_titles(), ["topic 1"])
        with Session(self.engine) as session:
            count = session.query(revise.Coverage).count()
        self.assertEqual(count, 1)

    def test_due_does_not_scan_markdown(self):
        tracker = self.make_tracker("# 12/06/2026\n\n- [x] new topic\n")
        due = revise.get_due_topics()
        self.assertEqual(due, [])
        revise.scan_tracker(self.make_config(tracker))
        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(revise.Coverage.title == "new topic").first()
        self.assertIsNotNone(topic)

    def test_old_date_is_ignored_even_if_it_has_x(self):
        tracker = self.make_tracker(
            "# 12/06/2026\n\n- [x] top topic\n\n# 13/06/2026\n\n- [x] newer topic\n- [x] another newer topic\n"
        )
        revise.scan_tracker(self.make_config(tracker))
        self.assertEqual(self.get_titles(), ["top topic"])


class BacklogTests(unittest.TestCase):
    """
    promote_overdue_to_backlog(), push_backlog_to_today(),
    get_backlog_topics(), and the due/backlog interaction in
    resolve_due_topics() -- the "missed revision -> backlog ->
    pushed back on a free day" feature.
    """

    @classmethod
    def setUpClass(cls):
        cls.engine = create_engine("sqlite:///:memory:", echo=False)
        revise.engine = cls.engine
        revise.Base.metadata.create_all(cls.engine)

    def setUp(self):
        with Session(self.engine) as session:
            for table in reversed(revise.Base.metadata.sorted_tables):
                session.execute(table.delete())
            session.commit()

    def add_coverage(self, title, covered_date=None, next_review=None,
                      interval_days=2, stage=0, status="active",
                      keywords=None, source_key=None):
        covered_date = covered_date or date.today()
        next_review = next_review or covered_date
        with Session(self.engine) as session:
            coverage = revise.Coverage(
                title=title,
                covered_date=covered_date,
                next_review=next_review,
                interval_days=interval_days,
                stage=stage,
                status=status,
                keywords=json.dumps(keywords or []),
                source_file="manual-test-entry",
                source_key=source_key or f"manual::{title}::{covered_date.isoformat()}"
            )
            session.add(coverage)
            session.commit()
            session.refresh(coverage)
            return coverage.id

    def get_status(self, coverage_id):
        with Session(self.engine) as session:
            return session.get(revise.Coverage, coverage_id).status

    def get_topic(self, coverage_id):
        with Session(self.engine) as session:
            return session.get(revise.Coverage, coverage_id)

    # ---- promote_overdue_to_backlog ----

    def test_promote_moves_only_strictly_overdue_active_topics(self):
        today = date.today()
        yesterday_id = self.add_coverage(
            "missed yesterday", next_review=today - timedelta(days=1)
        )
        today_id = self.add_coverage(
            "due today", next_review=today
        )
        future_id = self.add_coverage(
            "due tomorrow", next_review=today + timedelta(days=1)
        )

        moved = revise.promote_overdue_to_backlog()

        self.assertEqual(moved, 1)
        self.assertEqual(self.get_status(yesterday_id), "backlog")
        self.assertEqual(self.get_status(today_id), "active")
        self.assertEqual(self.get_status(future_id), "active")

    def test_promote_is_idempotent(self):
        today = date.today()
        topic_id = self.add_coverage(
            "missed", next_review=today - timedelta(days=3)
        )

        first = revise.promote_overdue_to_backlog()
        second = revise.promote_overdue_to_backlog()

        self.assertEqual(first, 1)
        self.assertEqual(second, 0)  # already backlog, nothing left to move
        self.assertEqual(self.get_status(topic_id), "backlog")

    def test_promote_ignores_removed_topics(self):
        today = date.today()
        topic_id = self.add_coverage(
            "removed but overdue",
            next_review=today - timedelta(days=5),
            status="removed"
        )

        revise.promote_overdue_to_backlog()

        self.assertEqual(self.get_status(topic_id), "removed")

    # ---- get_due_topics excludes backlog unless nothing else is due ----

    def test_due_excludes_backlog_when_something_else_is_due(self):
        today = date.today()
        self.add_coverage("missed", next_review=today - timedelta(days=2))
        due_today_id = self.add_coverage("due today", next_review=today)

        due_topics = revise.get_due_topics()

        self.assertEqual(len(due_topics), 1)
        self.assertEqual(due_topics[0].id, due_today_id)

    def test_due_pushes_backlog_when_nothing_else_is_due(self):
        today = date.today()
        missed_id = self.add_coverage(
            "missed", next_review=today - timedelta(days=4),
            interval_days=5, stage=2
        )

        due_topics = revise.get_due_topics()

        self.assertEqual(len(due_topics), 1)
        self.assertEqual(due_topics[0].id, missed_id)

        pushed = self.get_topic(missed_id)
        self.assertEqual(pushed.status, "active")
        self.assertEqual(pushed.next_review, today)
        # Interval/stage untouched by the push -- same review cycle.
        self.assertEqual(pushed.interval_days, 5)
        self.assertEqual(pushed.stage, 2)

    def test_due_push_is_idempotent_across_calls_same_day(self):
        today = date.today()
        self.add_coverage("missed", next_review=today - timedelta(days=1))

        first_call = revise.get_due_topics()
        second_call = revise.get_due_topics()

        self.assertEqual(len(first_call), 1)
        self.assertEqual(len(second_call), 1)  # not duplicated

    def test_due_push_respects_keyword_filter(self):
        today = date.today()
        dsa_id = self.add_coverage(
            "dsa missed", next_review=today - timedelta(days=1),
            keywords=["dsa"]
        )
        os_id = self.add_coverage(
            "os missed", next_review=today - timedelta(days=1),
            keywords=["os"]
        )

        due_topics = revise.get_due_topics(keyword="dsa")

        self.assertEqual(len(due_topics), 1)
        self.assertEqual(due_topics[0].id, dsa_id)

        # Only the matching one got pushed; the other stays in backlog.
        self.assertEqual(self.get_status(dsa_id), "active")
        self.assertEqual(self.get_status(os_id), "backlog")

    def test_due_with_no_topics_at_all_returns_empty(self):
        self.assertEqual(revise.get_due_topics(), [])

    # ---- get_backlog_topics never pushes ----

    def test_get_backlog_topics_does_not_push_into_due(self):
        today = date.today()
        missed_id = self.add_coverage(
            "missed", next_review=today - timedelta(days=3)
        )

        backlog_topics = revise.get_backlog_topics()

        self.assertEqual(len(backlog_topics), 1)
        self.assertEqual(backlog_topics[0].id, missed_id)
        # Still backlog -- inspecting it must not consume it.
        self.assertEqual(self.get_status(missed_id), "backlog")

    def test_get_backlog_topics_promotes_newly_overdue_first(self):
        today = date.today()
        topic_id = self.add_coverage(
            "just missed", next_review=today - timedelta(days=1)
        )

        # Not promoted yet -- still "active" in the DB.
        self.assertEqual(self.get_status(topic_id), "active")

        backlog_topics = revise.get_backlog_topics()

        self.assertEqual(len(backlog_topics), 1)
        self.assertEqual(self.get_status(topic_id), "backlog")

    def test_get_backlog_topics_ordered_oldest_first(self):
        today = date.today()
        newer_id = self.add_coverage(
            "missed 1 day", next_review=today - timedelta(days=1)
        )
        older_id = self.add_coverage(
            "missed 5 days", next_review=today - timedelta(days=5)
        )

        backlog_topics = revise.get_backlog_topics()

        self.assertEqual(
            [t.id for t in backlog_topics],
            [older_id, newer_id]
        )

    def test_get_backlog_topics_respects_keyword_filter(self):
        today = date.today()
        self.add_coverage(
            "dsa missed", next_review=today - timedelta(days=1),
            keywords=["dsa"]
        )
        self.add_coverage(
            "os missed", next_review=today - timedelta(days=1),
            keywords=["os"]
        )

        backlog_topics = revise.get_backlog_topics(keyword="os")

        self.assertEqual(len(backlog_topics), 1)
        self.assertEqual(backlog_topics[0].title, "os missed")

    def test_backlog_empty_when_nothing_missed(self):
        today = date.today()
        self.add_coverage("due today", next_review=today)

        self.assertEqual(revise.get_backlog_topics(), [])

    # ---- review_topics integrates with the push ----

    def test_review_pulls_from_backlog_when_nothing_due(self):
        today = date.today()
        missed_id = self.add_coverage(
            "missed", next_review=today - timedelta(days=2),
            interval_days=3
        )

        revise.review_topics("confident")

        topic = self.get_topic(missed_id)
        self.assertEqual(topic.interval_days, 5)  # 3 -> 5, confident
        self.assertEqual(topic.status, "active")
        self.assertEqual(topic.next_review, today + timedelta(days=5))

    # ---- show_due / show_backlog printed output ----

    def test_show_due_reports_backlog_pull(self):
        import io
        import contextlib

        today = date.today()
        self.add_coverage("missed", next_review=today - timedelta(days=1))

        captured = io.StringIO()
        with contextlib.redirect_stdout(captured):
            revise.show_due()

        self.assertIn("backlog", captured.getvalue().lower())

    def test_show_due_no_backlog_message_when_genuinely_due(self):
        import io
        import contextlib

        today = date.today()
        self.add_coverage("due today", next_review=today)

        captured = io.StringIO()
        with contextlib.redirect_stdout(captured):
            revise.show_due()

        self.assertNotIn("pulled", captured.getvalue().lower())

    def test_show_backlog_empty_message(self):
        import io
        import contextlib

        captured = io.StringIO()
        with contextlib.redirect_stdout(captured):
            revise.show_backlog()

        self.assertIn("Backlog is empty", captured.getvalue())

    def test_show_backlog_lists_missed_topics(self):
        import io
        import contextlib

        today = date.today()
        self.add_coverage("missed topic", next_review=today - timedelta(days=3))

        captured = io.StringIO()
        with contextlib.redirect_stdout(captured):
            revise.show_backlog()

        output = captured.getvalue()
        self.assertIn("missed topic", output)
        self.assertIn("Missed by: 3 day(s)", output)

    # ---- list_topics shows accurate backlog status ----

    def test_list_topics_shows_backlog_status_after_promotion(self):
        import io
        import contextlib

        today = date.today()
        self.add_coverage("missed", next_review=today - timedelta(days=2))

        captured = io.StringIO()
        with contextlib.redirect_stdout(captured):
            revise.list_topics()

        self.assertIn("backlog", captured.getvalue())

    # ---- remove_topics can remove backlog items ----

    def test_remove_topics_can_remove_a_backlog_item(self):
        import builtins

        today = date.today()
        topic_id = self.add_coverage(
            "missed", next_review=today - timedelta(days=2)
        )
        revise.promote_overdue_to_backlog()
        self.assertEqual(self.get_status(topic_id), "backlog")

        original_input = builtins.input
        builtins.input = lambda prompt="": "y"
        try:
            revise.remove_topics(title="missed")
        finally:
            builtins.input = original_input

        self.assertEqual(self.get_status(topic_id), "removed")


class ReviewSchedulingTests(unittest.TestCase):
    """
    calculate_next_interval() and review_topics() implement the
    actual spaced-repetition schedule. This is the core value of the
    app and previously had zero test coverage.
    """

    @classmethod
    def setUpClass(cls):
        cls.engine = create_engine("sqlite:///:memory:", echo=False)
        revise.engine = cls.engine
        revise.Base.metadata.create_all(cls.engine)

    def setUp(self):
        with Session(self.engine) as session:
            for table in reversed(revise.Base.metadata.sorted_tables):
                session.execute(table.delete())
            session.commit()

    def add_coverage(self, title, covered_date=None, next_review=None,
                      interval_days=2, stage=0, status="active",
                      keywords=None, source_key=None):
        import json
        covered_date = covered_date or date.today()
        next_review = next_review or covered_date
        with Session(self.engine) as session:
            coverage = revise.Coverage(
                title=title,
                covered_date=covered_date,
                next_review=next_review,
                interval_days=interval_days,
                stage=stage,
                status=status,
                keywords=json.dumps(keywords or []),
                source_file="manual-test-entry",
                source_key=source_key or f"manual::{title}::{covered_date.isoformat()}"
            )
            session.add(coverage)
            session.commit()
            session.refresh(coverage)
            return coverage.id

    # ---- calculate_next_interval: pure function, no DB needed ----

    def test_first_pass_schedule_confident(self):
        # 2 -> 3 -> 5 -> 7 -> 14 (after 7, the schedule doubles)
        self.assertEqual(revise.calculate_next_interval(2, "confident"), 3)
        self.assertEqual(revise.calculate_next_interval(3, "confident"), 5)
        self.assertEqual(revise.calculate_next_interval(5, "confident"), 7)
        self.assertEqual(revise.calculate_next_interval(7, "confident"), 14)

    def test_second_pass_schedule_doubles(self):
        # 14 -> 28 -> 56
        self.assertEqual(revise.calculate_next_interval(14, "confident"), 28)
        self.assertEqual(revise.calculate_next_interval(28, "confident"), 56)

    def test_failed_always_resets_to_2(self):
        for interval in (2, 3, 5, 7, 14, 56, 200):
            self.assertEqual(
                revise.calculate_next_interval(interval, "failed"), 2
            )

    def test_invalid_result_raises(self):
        with self.assertRaises(ValueError):
            revise.calculate_next_interval(5, "not_a_real_result")

    # ---- review_topics: DB side effects ----

    def test_confident_review_advances_interval_and_stage(self):
        today = date.today()
        coverage_id = self.add_coverage(
            "topic a", covered_date=today - timedelta(days=5),
            next_review=today - timedelta(days=1),
            interval_days=3, stage=1
        )

        revise.review_topics("confident")

        with Session(self.engine) as session:
            topic = session.get(revise.Coverage, coverage_id)

        self.assertEqual(topic.interval_days, 5)  # 3 -> 5 in first-pass schedule
        self.assertEqual(topic.stage, 2)           # stage increments on confident
        self.assertEqual(topic.next_review, today + timedelta(days=5))

    def test_failed_review_resets_interval_and_stage(self):
        today = date.today()
        coverage_id = self.add_coverage(
            "topic b", covered_date=today - timedelta(days=20),
            next_review=today,
            interval_days=14, stage=4
        )

        revise.review_topics("failed")

        with Session(self.engine) as session:
            topic = session.get(revise.Coverage, coverage_id)

        self.assertEqual(topic.interval_days, 2)
        self.assertEqual(topic.stage, 0)
        self.assertEqual(topic.next_review, today + timedelta(days=2))

    def test_review_writes_history_row(self):
        today = date.today()
        self.add_coverage(
            "topic c", covered_date=today, next_review=today, interval_days=2
        )

        revise.review_topics("confident")

        with Session(self.engine) as session:
            history = session.query(revise.ReviewHistory).all()

        self.assertEqual(len(history), 1)
        self.assertEqual(history[0].result, "confident")
        self.assertEqual(history[0].previous_interval, 2)
        self.assertEqual(history[0].new_interval, 3)

    def test_review_does_not_touch_not_yet_due_topics(self):
        today = date.today()
        due_id = self.add_coverage(
            "due topic", covered_date=today, next_review=today, interval_days=2
        )
        not_due_id = self.add_coverage(
            "future topic", covered_date=today,
            next_review=today + timedelta(days=10), interval_days=2
        )

        revise.review_topics("confident")

        with Session(self.engine) as session:
            due_topic = session.get(revise.Coverage, due_id)
            future_topic = session.get(revise.Coverage, not_due_id)

        self.assertEqual(due_topic.interval_days, 3)
        self.assertEqual(future_topic.interval_days, 2)  # untouched

    def test_review_respects_keyword_filter(self):
        today = date.today()
        dsa_id = self.add_coverage(
            "dsa topic", covered_date=today, next_review=today,
            interval_days=2, keywords=["dsa"]
        )
        os_id = self.add_coverage(
            "os topic", covered_date=today, next_review=today,
            interval_days=2, keywords=["os"]
        )

        revise.review_topics("confident", keyword="dsa")

        with Session(self.engine) as session:
            dsa_topic = session.get(revise.Coverage, dsa_id)
            os_topic = session.get(revise.Coverage, os_id)

        self.assertEqual(dsa_topic.interval_days, 3)   # reviewed
        self.assertEqual(os_topic.interval_days, 2)    # untouched

    def test_review_with_no_due_topics_is_a_no_op(self):
        today = date.today()
        self.add_coverage(
            "future topic", covered_date=today,
            next_review=today + timedelta(days=5)
        )

        # Should not raise, should not touch anything.
        revise.review_topics("confident")

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).first()

        self.assertEqual(topic.interval_days, 2)


class KeywordAndTitleParsingTests(unittest.TestCase):
    """
    extract_bold_keywords(), clean_topic_title() and the nested-bullet
    hierarchy logic in parse_tracker() had no direct coverage before.
    """

    def make_tracker(self, content):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "tracker.md"
        path.write_text(content, encoding="utf-8")
        return path

    def test_extract_bold_keywords_basic(self):
        self.assertEqual(
            revise.extract_bold_keywords("Binary Search **dsa** **array**"),
            ["dsa", "array"]
        )

    def test_extract_bold_keywords_lowercases(self):
        self.assertEqual(
            revise.extract_bold_keywords("Topic **DSA**"),
            ["dsa"]
        )

    def test_extract_bold_keywords_none_present(self):
        self.assertEqual(
            revise.extract_bold_keywords("Plain topic, no keywords"),
            []
        )

    def test_clean_topic_title_keeps_link_text_strips_brackets_and_url(self):
        self.assertEqual(
            revise.clean_topic_title(
                "[Binary Search](./binary-search.md) **dsa**"
            ),
            "Binary Search"
        )

    def test_clean_topic_title_link_as_entire_topic(self):
        """
        Regression test: previously the whole [text](url) match was
        replaced with nothing, so a bullet that was ENTIRELY a link
        (e.g. "- [dsa problem](link)") collapsed to an empty title
        and got silently dropped. Only the brackets and the "(link)"
        URL should be discarded -- the visible text is the topic.
        """

        self.assertEqual(
            revise.clean_topic_title("[dsa problem](link)"),
            "dsa problem"
        )

    def test_clean_topic_title_strips_angle_links(self):
        self.assertEqual(
            revise.clean_topic_title(
                "Read the RFC <https://example.com/rfc>"
            ),
            "Read the RFC"
        )

    def test_sibling_bold_keywords_do_not_chain(self):
        """
        Documents ACTUAL behavior, which differs from print_template()'s
        docstring. The template shows:

            - Binary Search **dsa**
                - **array**
                - **hard**

        and claims this produces "dsa/array/hard". It does not: "array"
        and "hard" are siblings (same indentation under "Binary Search"),
        so each becomes its own branch off "dsa" rather than a chain.
        """

        tracker = self.make_tracker(
            "17/08/2026\n\n"
            "- Binary Search **dsa**\n"
            "    - **array**\n"
            "    - **hard**\n"
        )

        results = revise.parse_tracker(tracker)

        self.assertEqual(len(results), 1)
        self.assertEqual(
            results[0]["keywords"],
            ["dsa", "dsa/array", "dsa/hard"]
        )

    def test_true_nesting_does_chain(self):
        """
        To actually get "dsa/array/hard", "hard" must be indented
        deeper than "array" (true nesting), not at the same level.
        """

        tracker = self.make_tracker(
            "17/08/2026\n\n"
            "- Binary Search **dsa**\n"
            "    - **array**\n"
            "        - **hard**\n"
        )

        results = revise.parse_tracker(tracker)

        self.assertEqual(
            results[0]["keywords"],
            ["dsa", "dsa/array", "dsa/array/hard"]
        )

    # ---- slash-shorthand for hierarchical tags ----

    def test_expand_keyword_path_basic(self):
        self.assertEqual(
            revise.expand_keyword_path("dsa/array/hard", ""),
            ["dsa", "dsa/array", "dsa/array/hard"]
        )

    def test_expand_keyword_path_with_parent(self):
        self.assertEqual(
            revise.expand_keyword_path("array/hard", "dsa"),
            ["dsa/array", "dsa/array/hard"]
        )

    def test_expand_keyword_path_no_slash(self):
        self.assertEqual(
            revise.expand_keyword_path("dsa", ""),
            ["dsa"]
        )

    def test_expand_keyword_path_drops_empty_segments(self):
        # A stray double slash shouldn't produce an empty path segment.
        self.assertEqual(
            revise.expand_keyword_path("dsa//array", ""),
            ["dsa", "dsa/array"]
        )

    def test_slash_shorthand_top_level_matches_manual_nesting(self):
        tracker = self.make_tracker(
            "17/08/2026\n\n"
            "- Binary Search **dsa/array/hard**\n"
        )

        results = revise.parse_tracker(tracker)

        self.assertEqual(
            results[0]["keywords"],
            ["dsa", "dsa/array", "dsa/array/hard"]
        )

    def test_slash_shorthand_nested_under_parent(self):
        tracker = self.make_tracker(
            "17/08/2026\n\n"
            "- Binary Search **dsa**\n"
            "    - **array/hard**\n"
        )

        results = revise.parse_tracker(tracker)

        self.assertEqual(
            results[0]["keywords"],
            ["dsa", "dsa/array", "dsa/array/hard"]
        )

    def test_slash_shorthand_continues_nesting_for_later_bullets(self):
        """
        A nested bullet placed after a slash-shorthand tag should
        continue nesting under the shorthand's deepest segment.
        """

        tracker = self.make_tracker(
            "17/08/2026\n\n"
            "- Binary Search **dsa/array**\n"
            "    - **hard**\n"
        )

        results = revise.parse_tracker(tracker)

        self.assertEqual(
            results[0]["keywords"],
            ["dsa", "dsa/array", "dsa/array/hard"]
        )

    def test_slash_shorthand_mixed_with_plain_root_keyword(self):
        """
        With multiple bold tags on the root bullet, nested bullets
        attach under the LAST tag (unchanged prior behavior) -- here
        the last tag is a plain keyword, not the slash path before it.
        """

        tracker = self.make_tracker(
            "17/08/2026\n\n"
            "- Binary Search **dsa/array** **os**\n"
            "    - **scheduling**\n"
        )

        results = revise.parse_tracker(tracker)

        self.assertEqual(
            results[0]["keywords"],
            ["dsa", "dsa/array", "os", "os/scheduling"]
        )

    def test_slash_shorthand_filterable_at_every_level(self):
        tracker = self.make_tracker(
            "17/08/2026\n\n"
            "- Binary Search **dsa/array/hard**\n"
        )

        results = revise.parse_tracker(tracker)
        keywords = results[0]["keywords"]

        self.assertIn("dsa", keywords)
        self.assertIn("dsa/array", keywords)
        self.assertIn("dsa/array/hard", keywords)

    def test_markdown_link_text_kept_url_stripped_keyword_kept(self):
        tracker = self.make_tracker(
            "17/08/2026\n\n"
            "- [Binary Search](./x.md) **dsa**\n"
        )

        results = revise.parse_tracker(tracker)

        self.assertEqual(results[0]["title"], "Binary Search")
        self.assertEqual(results[0]["keywords"], ["dsa"])

    def test_bullet_that_is_entirely_a_link_is_not_dropped(self):
        """
        Regression test at the parse_tracker level: a bullet whose
        entire content is a Markdown link must still produce a topic
        using the link's visible text, not be silently skipped for
        having an "empty" title.
        """

        tracker = self.make_tracker(
            "17/08/2026\n\n"
            "- [dsa problem](https://leetcode.com/problems/two-sum) **dsa**\n"
        )

        results = revise.parse_tracker(tracker)

        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["title"], "dsa problem")
        self.assertEqual(results[0]["keywords"], ["dsa"])


class FilterAndDateRangeTests(unittest.TestCase):
    """
    topic_matches_filter() (hierarchical keyword/title matching) and
    parse_date_filter() (single date / "::" ranges) had no direct
    coverage before.
    """

    class FakeTopic:
        def __init__(self, title, keywords, covered_date):
            import json
            self.title = title
            self.keywords = json.dumps(keywords)
            self.covered_date = covered_date

    def test_filter_matches_title_substring(self):
        topic = self.FakeTopic("Binary Search", [], date.today())
        self.assertTrue(revise.topic_matches_filter(topic, "binary"))

    def test_filter_matches_exact_hierarchical_path(self):
        topic = self.FakeTopic(
            "Binary Search", ["dsa", "dsa/array", "dsa/array/hard"], date.today()
        )
        self.assertTrue(revise.topic_matches_filter(topic, "dsa/array"))

    def test_filter_matches_individual_keyword_segment(self):
        topic = self.FakeTopic(
            "Binary Search", ["dsa/array/hard"], date.today()
        )
        # "array" alone should match even though only the full path is stored.
        self.assertTrue(revise.topic_matches_filter(topic, "array"))

    def test_filter_no_match_returns_false(self):
        topic = self.FakeTopic("Binary Search", ["dsa"], date.today())
        self.assertFalse(revise.topic_matches_filter(topic, "dbms"))

    def test_empty_filter_matches_everything(self):
        topic = self.FakeTopic("Binary Search", [], date.today())
        self.assertTrue(revise.topic_matches_filter(topic, None))
        self.assertTrue(revise.topic_matches_filter(topic, ""))

    def test_parse_date_filter_single_date(self):
        start, end = revise.parse_date_filter("13/05/2025")
        self.assertEqual(start, date(2025, 5, 13))
        self.assertEqual(end, date(2025, 5, 13))

    def test_parse_date_filter_full_range(self):
        start, end = revise.parse_date_filter("13/05/2025::24/07/2026")
        self.assertEqual(start, date(2025, 5, 13))
        self.assertEqual(end, date(2026, 7, 24))

    def test_parse_date_filter_open_start(self):
        start, end = revise.parse_date_filter("::24/07/2026")
        self.assertIsNone(start)
        self.assertEqual(end, date(2026, 7, 24))

    def test_parse_date_filter_open_end(self):
        start, end = revise.parse_date_filter("13/05/2025::")
        self.assertEqual(start, date(2025, 5, 13))
        self.assertIsNone(end)

    def test_parse_date_filter_rejects_reversed_range(self):
        with self.assertRaises(ValueError):
            revise.parse_date_filter("24/07/2026::13/05/2025")

    def test_parse_date_filter_rejects_malformed_date(self):
        with self.assertRaises(ValueError):
            revise.parse_date_filter("13-05-2025")

    def test_parse_date_filter_rejects_impossible_date(self):
        # Not a real calendar date.
        with self.assertRaises(ValueError):
            revise.parse_date_filter("31/13/2026")


class MalformedInputTests(unittest.TestCase):
    """
    parse_tracker() should degrade gracefully on malformed lines
    rather than crashing or silently misattributing bullets.
    """

    def make_tracker(self, content):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "tracker.md"
        path.write_text(content, encoding="utf-8")
        return path

    def test_invalid_calendar_date_is_treated_as_no_date(self):
        """
        "31/13/2026" matches DATE_PATTERN's shape (DD/MM/YYYY) but
        isn't a real date, so parse_date() returns None. Any bullets
        that follow should be skipped (no active date), not crash.
        """

        tracker = self.make_tracker(
            "31/13/2026\n\n- some topic\n"
        )

        results = revise.parse_tracker(tracker)
        self.assertEqual(results, [])

    def test_bullets_before_any_date_are_ignored(self):
        tracker = self.make_tracker(
            "- orphan topic before any date\n\n12/06/2026\n\n- real topic\n"
        )

        results = revise.parse_tracker(tracker)
        self.assertEqual([r["title"] for r in results], ["real topic"])

    def test_missing_tracker_file_raises_file_not_found(self):
        with self.assertRaises(FileNotFoundError):
            revise.parse_tracker("/no/such/path/tracker.md")

    def test_checkbox_with_invalid_marker_is_not_treated_as_checkbox(self):
        # "[y]" isn't " " or "x"/"X", so CHECKBOX_PATTERN won't match it;
        # it should be kept as literal text in the title instead.
        tracker = self.make_tracker(
            "12/06/2026\n\n- [y] weird topic\n"
        )

        results = revise.parse_tracker(tracker)
        self.assertEqual(len(results), 1)
        self.assertIsNone(results[0]["checked"])
        self.assertIn("weird topic", results[0]["title"])


class ExportExcelTests(unittest.TestCase):
    """
    export_excel() had zero coverage before. It writes to
    DATA_DIR / "revision_export.xlsx" -- DATA_DIR is a module-level
    global, so tests monkeypatch it to a temp directory rather than
    touching the real data folder.
    """

    @classmethod
    def setUpClass(cls):
        cls.engine = create_engine("sqlite:///:memory:", echo=False)
        revise.engine = cls.engine
        revise.Base.metadata.create_all(cls.engine)

    def setUp(self):
        with Session(self.engine) as session:
            for table in reversed(revise.Base.metadata.sorted_tables):
                session.execute(table.delete())
            session.commit()

        self._original_data_dir = revise.DATA_DIR
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        self.addCleanup(self._restore_data_dir)
        revise.DATA_DIR = Path(temp_dir.name)

    def _restore_data_dir(self):
        revise.DATA_DIR = self._original_data_dir

    def add_coverage(self, title, covered_date=None, next_review=None,
                      interval_days=2, stage=0, status="active",
                      keywords=None, source_file="tracker.md", source_key=None):
        import json
        covered_date = covered_date or date.today()
        next_review = next_review or covered_date
        with Session(self.engine) as session:
            coverage = revise.Coverage(
                title=title,
                covered_date=covered_date,
                next_review=next_review,
                interval_days=interval_days,
                stage=stage,
                status=status,
                keywords=json.dumps(keywords or []),
                source_file=source_file,
                source_key=source_key or f"manual::{title}::{covered_date.isoformat()}"
            )
            session.add(coverage)
            session.commit()

    def test_export_creates_file_with_correct_headers(self):
        self.add_coverage("Binary Search", keywords=["dsa", "dsa/array"])

        revise.export_excel()

        output_path = revise.DATA_DIR / "revision_export.xlsx"
        self.assertTrue(output_path.exists())

        from openpyxl import load_workbook
        workbook = load_workbook(output_path)
        sheet = workbook.active

        self.assertEqual(sheet.title, "All Topics")
        header_row = [cell.value for cell in sheet[1]]
        self.assertEqual(
            header_row,
            ["ID", "Date", "Topic", "Keywords", "Next Review",
             "Interval", "Stage", "Status", "Source File"]
        )

    def test_export_row_values_match_topic_fields(self):
        covered = date(2026, 6, 12)
        due = date(2026, 6, 14)
        self.add_coverage(
            "Binary Search", covered_date=covered, next_review=due,
            interval_days=2, stage=1, status="active",
            keywords=["dsa", "dsa/array"], source_file="/path/tracker.md"
        )

        revise.export_excel()

        from openpyxl import load_workbook
        output_path = revise.DATA_DIR / "revision_export.xlsx"
        sheet = load_workbook(output_path).active

        data_row = [cell.value for cell in sheet[2]]

        self.assertEqual(data_row[1], "12/06/2026")           # Date
        self.assertEqual(data_row[2], "Binary Search")        # Topic
        self.assertEqual(data_row[3], "dsa | dsa/array")      # Keywords
        self.assertEqual(data_row[4], "14/06/2026")           # Next Review
        self.assertEqual(data_row[5], 2)                      # Interval
        self.assertEqual(data_row[6], 1)                      # Stage
        self.assertEqual(data_row[7], "active")                # Status
        self.assertEqual(data_row[8], "/path/tracker.md")     # Source File

    def test_export_includes_removed_topics_without_filtering(self):
        """
        Documents actual behavior: export_excel(), like list_topics(),
        does not filter by status, so 'removed' topics are still
        included in the export (with their status visible in the
        Status column) rather than being excluded.
        """

        self.add_coverage("Active Topic", status="active")
        self.add_coverage("Removed Topic", status="removed")

        revise.export_excel()

        from openpyxl import load_workbook
        output_path = revise.DATA_DIR / "revision_export.xlsx"
        sheet = load_workbook(output_path).active

        titles = [row[2] for row in sheet.iter_rows(min_row=2, values_only=True)]
        self.assertIn("Active Topic", titles)
        self.assertIn("Removed Topic", titles)

    def test_export_with_no_topics_still_writes_header_only_file(self):
        revise.export_excel()

        from openpyxl import load_workbook
        output_path = revise.DATA_DIR / "revision_export.xlsx"
        self.assertTrue(output_path.exists())

        sheet = load_workbook(output_path).active
        self.assertEqual(sheet.max_row, 1)  # header row only

    def test_export_overwrites_previous_file(self):
        self.add_coverage("First Export Topic")
        revise.export_excel()

        with Session(self.engine) as session:
            for table in reversed(revise.Base.metadata.sorted_tables):
                session.execute(table.delete())
            session.commit()

        self.add_coverage("Second Export Topic")
        revise.export_excel()

        from openpyxl import load_workbook
        output_path = revise.DATA_DIR / "revision_export.xlsx"
        sheet = load_workbook(output_path).active

        titles = [row[2] for row in sheet.iter_rows(min_row=2, values_only=True)]
        self.assertEqual(titles, ["Second Export Topic"])

    def test_export_missing_openpyxl_prints_friendly_message_and_does_not_crash(self):
        """
        If openpyxl isn't installed, export_excel() should print
        install instructions and return quietly rather than raising.
        """

        import builtins
        import io
        import contextlib

        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("No module named 'openpyxl'")
            return real_import(name, *args, **kwargs)

        captured = io.StringIO()

        builtins.__import__ = fake_import
        try:
            with contextlib.redirect_stdout(captured):
                revise.export_excel()  # should not raise
        finally:
            builtins.__import__ = real_import

        output_path = revise.DATA_DIR / "revision_export.xlsx"
        self.assertFalse(output_path.exists())
        self.assertIn("pip install openpyxl", captured.getvalue())


class RemoveTopicsTests(unittest.TestCase):
    """
    remove_topics() had no coverage before. It's destructive (marks
    rows as 'removed'), so it's worth pinning down precisely.
    """

    @classmethod
    def setUpClass(cls):
        cls.engine = create_engine("sqlite:///:memory:", echo=False)
        revise.engine = cls.engine
        revise.Base.metadata.create_all(cls.engine)

    def setUp(self):
        with Session(self.engine) as session:
            for table in reversed(revise.Base.metadata.sorted_tables):
                session.execute(table.delete())
            session.commit()

    def add_coverage(self, title, covered_date=None, keywords=None, source_key=None):
        import json
        covered_date = covered_date or date.today()
        with Session(self.engine) as session:
            coverage = revise.Coverage(
                title=title,
                covered_date=covered_date,
                next_review=covered_date,
                interval_days=2,
                stage=0,
                status="active",
                keywords=json.dumps(keywords or []),
                source_file="manual-test-entry",
                source_key=source_key or f"manual::{title}::{covered_date.isoformat()}"
            )
            session.add(coverage)
            session.commit()
            session.refresh(coverage)
            return coverage.id

    def test_remove_by_title_marks_removed_on_confirmation(self, monkeypatch=None):
        import builtins
        coverage_id = self.add_coverage("Binary Search")

        original_input = builtins.input
        builtins.input = lambda prompt="": "y"
        try:
            revise.remove_topics(title="Binary")
        finally:
            builtins.input = original_input

        with Session(self.engine) as session:
            topic = session.get(revise.Coverage, coverage_id)
        self.assertEqual(topic.status, "removed")

    def test_remove_cancelled_leaves_topic_active(self):
        import builtins
        coverage_id = self.add_coverage("Binary Search")

        original_input = builtins.input
        builtins.input = lambda prompt="": "n"
        try:
            revise.remove_topics(title="Binary")
        finally:
            builtins.input = original_input

        with Session(self.engine) as session:
            topic = session.get(revise.Coverage, coverage_id)
        self.assertEqual(topic.status, "active")

    def test_remove_does_not_match_unrelated_titles(self):
        import builtins
        keep_id = self.add_coverage("Process Scheduling")
        self.add_coverage("Binary Search")

        original_input = builtins.input
        builtins.input = lambda prompt="": "y"
        try:
            revise.remove_topics(title="Binary")
        finally:
            builtins.input = original_input

        with Session(self.engine) as session:
            kept = session.get(revise.Coverage, keep_id)
        self.assertEqual(kept.status, "active")

    def test_removed_topic_reactivates_on_rescan(self):
        """
        import_topics() should flip a 'removed' record back to
        'active' if the same topic is checked again in a later scan.
        """

        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        tracker_path = Path(temp_dir.name) / "tracker.md"
        tracker_path.write_text("12/06/2026\n\n- [x] revive me\n", encoding="utf-8")

        config = {"name": "Test User", "tracker_path": str(tracker_path), "last_scan": None}

        revise.scan_tracker(config)

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(
                revise.Coverage.title == "revive me"
            ).first()
            topic_id = topic.id
            topic.status = "removed"
            session.commit()

        # Re-scanning the same checked topic should reactivate it.
        revise.scan_tracker(config)

        with Session(self.engine) as session:
            topic = session.get(revise.Coverage, topic_id)
        self.assertEqual(topic.status, "active")


class CliRunnerTests(unittest.TestCase):
    """
    Exercises the Typer CLI plumbing itself (revise.app) rather than
    calling the underlying functions directly -- this is the surface
    that changed with the argparse -> Typer rewrite, and it wasn't
    covered by any of the tests above.
    """

    @classmethod
    def setUpClass(cls):
        cls.engine = create_engine("sqlite:///:memory:", echo=False)
        revise.engine = cls.engine
        revise.Base.metadata.create_all(cls.engine)
        cls.runner = CliRunner()

    def setUp(self):
        with Session(self.engine) as session:
            for table in reversed(revise.Base.metadata.sorted_tables):
                session.execute(table.delete())
            session.commit()

        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        self.temp_path = Path(temp_dir.name)

        self.tracker_path = self.temp_path / "tracker.md"
        self.tracker_path.write_text(
            "12/06/2026\n\n- [x] cli topic **dsa/array**\n",
            encoding="utf-8"
        )

        self.config_path = self.temp_path / "config.json"
        self.config_path.write_text(
            json.dumps({
                "name": "Test User",
                "tracker_path": str(self.tracker_path),
                "last_scan": None
            }),
            encoding="utf-8"
        )

        self._original_config_file = revise.CONFIG_FILE
        self._original_data_dir = revise.DATA_DIR
        revise.CONFIG_FILE = self.config_path
        revise.DATA_DIR = self.temp_path
        self.addCleanup(self._restore_globals)

    def _restore_globals(self):
        revise.CONFIG_FILE = self._original_config_file
        revise.DATA_DIR = self._original_data_dir

    def invoke(self, args, input=None):
        return self.runner.invoke(revise.app, args, input=input)

    def test_help_does_not_prompt_for_setup(self):
        """
        Regression test for the argparse version's behavior, where
        --help ran the interactive setup() wizard before parsing args.
        Click resolves --help eagerly, before the app callback runs,
        so this should exit cleanly with no prompt.
        """

        result = self.invoke(["--help"])

        self.assertEqual(result.exit_code, 0)
        self.assertIn("Usage", result.output)
        self.assertIn("scan", result.output)
        self.assertIn("review", result.output)

    def test_scan_via_cli_imports_topic(self):
        result = self.invoke(["scan"])

        self.assertEqual(result.exit_code, 0)
        self.assertIn("New: 1", result.output)

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(
                revise.Coverage.title == "cli topic"
            ).first()

        self.assertIsNotNone(topic)

    def test_list_via_cli_scans_then_prints(self):
        result = self.invoke(["list"])

        self.assertEqual(result.exit_code, 0)
        self.assertIn("cli topic", result.output)

    def test_due_via_cli_reads_db_without_scanning(self):
        # Without scanning first, "due" should find nothing.
        result = self.invoke(["due"])
        self.assertIn("Nothing is due", result.output)

        self.invoke(["scan"])

        result = self.invoke(["due"])
        self.assertIn("cli topic", result.output)

    def test_review_confident_via_cli_updates_interval(self):
        self.invoke(["scan"])

        result = self.invoke(["review", "confident"])
        self.assertEqual(result.exit_code, 0)
        self.assertIn("Reviewed", result.output)

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(
                revise.Coverage.title == "cli topic"
            ).first()

        self.assertEqual(topic.interval_days, 3)  # 2 -> 3, confident

    def test_review_rejects_invalid_result_with_nonzero_exit(self):
        result = self.invoke(["review", "not-a-real-result"])

        self.assertNotEqual(result.exit_code, 0)
        self.assertIn(
            "confident",
            result.output.lower() + str(result.exception)
        )

    def test_review_filter_option_via_cli(self):
        self.invoke(["scan"])

        result = self.invoke(["review", "confident", "--filter", "dsa"])
        self.assertEqual(result.exit_code, 0)
        self.assertIn("Reviewed 1", result.output)

    def test_config_no_subcommand_shows_usage_hint(self):
        result = self.invoke(["config"])

        self.assertIn("revise config update", result.output)

    def test_config_update_via_cli_writes_config_file(self):
        result = self.invoke(
            ["config", "update", "--name", "New Name"]
        )

        self.assertEqual(result.exit_code, 0)

        updated = json.loads(self.config_path.read_text())
        self.assertEqual(updated["name"], "New Name")
        # Path untouched since --path wasn't passed.
        self.assertEqual(updated["tracker_path"], str(self.tracker_path))

    def test_config_update_with_no_options_prints_example(self):
        result = self.invoke(["config", "update"])

        self.assertIn("Example:", result.output)

    def test_template_via_cli_prints_guide(self):
        result = self.invoke(["template"])

        self.assertEqual(result.exit_code, 0)
        self.assertIn("Basic tracker format", result.output)

    def test_export_via_cli_writes_xlsx(self):
        self.invoke(["scan"])

        result = self.invoke(["export"])

        self.assertEqual(result.exit_code, 0)
        self.assertTrue(
            (self.temp_path / "revision_export.xlsx").exists()
        )

    def test_list_export_flag_also_writes_xlsx(self):
        result = self.invoke(["list", "--export"])

        self.assertEqual(result.exit_code, 0)
        self.assertTrue(
            (self.temp_path / "revision_export.xlsx").exists()
        )

    def test_remove_declined_via_cli_leaves_topic_active(self):
        self.invoke(["scan"])

        result = self.invoke(["remove", "cli"], input="n\n")

        self.assertEqual(result.exit_code, 0)
        self.assertIn("Cancelled", result.output)

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(
                revise.Coverage.title == "cli topic"
            ).first()

        self.assertEqual(topic.status, "active")

    def test_remove_confirmed_via_cli_marks_removed(self):
        self.invoke(["scan"])

        result = self.invoke(["remove", "cli"], input="y\n")

        self.assertEqual(result.exit_code, 0)

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(
                revise.Coverage.title == "cli topic"
            ).first()

        self.assertEqual(topic.status, "removed")

    def test_remove_invalid_date_filter_reports_error_not_crash(self):
        self.invoke(["scan"])

        result = self.invoke(["remove", "cli", "--date", "not-a-date"])

        self.assertEqual(result.exit_code, 0)  # handled, not raised
        self.assertIn("Invalid date filter", result.output)

    def test_no_command_shows_help(self):
        result = self.invoke([])

        self.assertEqual(result.exit_code, 0)
        self.assertIn("Usage", result.output)

    def test_backlog_command_shows_missed_topic(self):
        self.invoke(["scan"])

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(
                revise.Coverage.title == "cli topic"
            ).first()
            topic.next_review = date.today() - timedelta(days=3)
            session.commit()

        result = self.invoke(["backlog"])

        self.assertEqual(result.exit_code, 0)
        self.assertIn("cli topic", result.output)
        self.assertIn("Missed by: 3 day(s)", result.output)

    def test_backlog_command_empty_message(self):
        result = self.invoke(["backlog"])

        self.assertEqual(result.exit_code, 0)
        self.assertIn("Backlog is empty", result.output)

    def test_backlog_command_does_not_consume_backlog(self):
        self.invoke(["scan"])

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(
                revise.Coverage.title == "cli topic"
            ).first()
            topic.next_review = date.today() - timedelta(days=2)
            session.commit()

        self.invoke(["backlog"])  # inspect, should not push

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(
                revise.Coverage.title == "cli topic"
            ).first()

        self.assertEqual(topic.status, "backlog")

    def test_due_command_pushes_backlog_when_nothing_else_due(self):
        self.invoke(["scan"])

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(
                revise.Coverage.title == "cli topic"
            ).first()
            topic.next_review = date.today() - timedelta(days=2)
            session.commit()

        result = self.invoke(["due"])

        self.assertEqual(result.exit_code, 0)
        self.assertIn("cli topic", result.output)
        self.assertIn("backlog", result.output.lower())

        with Session(self.engine) as session:
            topic = session.query(revise.Coverage).filter(
                revise.Coverage.title == "cli topic"
            ).first()

        self.assertEqual(topic.status, "active")
        self.assertEqual(topic.next_review, date.today())


if __name__ == "__main__":
    unittest.main(verbosity=2)